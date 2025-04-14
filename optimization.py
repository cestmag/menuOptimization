def get_optimized_menu(mode, menu_id=None):
    """
    CSVとExcelからメニュー最適化を複数回試行し、
    各試行の結果を { "trial_1": {料理名: {quantity, image_url, price}, ... },
                     "trial_2": { ... }, ... } の形式で返す関数。

    Parameters:
        mode (str): '普通', 'タンパク質', 'カロリー'
        url (str): セキュリティチェック用（未使用）

    Returns:
        dict: 各試行ごとの最適化結果をまとめた連想配列
    """
    import csv
    import pulp
    import openpyxl
    import random
    import time

    # ★ ファイルパスの設定（適宜パスを変更してください）
    menu_csv_file = "menu_csv/" + menu_id + ".csv"    # プログラム1で作成されたCSVファイル
    std_excel_file = "data_std.xlsx"       # Excel形式の基準値ファイル

    # ★ CSVからのデータ読み込み
    menus = []         # 料理名リスト
    prices = {}        # 料理ごとの価格
    nutrients = {}     # 料理ごとの栄養素データ（数値リスト）
    image_urls = {}    # 料理ごとの画像URL

    with open(menu_csv_file, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        data = list(reader)

    # ヘッダー例: ["料理名", "価格", "エネルギー", "タンパク質", ..., "ビタミンC", "画像URL"]
    # 栄養素は「料理名」と「価格」、および「画像URL」を除いた部分
    header = data[0]
    nutrient_names = header[2:-1]  # 最後の列(画像URL)は除外

    for row in data[1:]:
        if not row:
            continue
        menu_name = row[0]
        price = int(row[1])
        nutrient_values = [float(v) if v else 0.0 for v in row[2:-1]]
        image_url = row[-1]

        menus.append(menu_name)
        prices[menu_name] = price
        nutrients[menu_name] = nutrient_values
        image_urls[menu_name] = image_url

    # ★ Excelから基準値読み込み（B列：必要量、C列：最大値）
    required_nutrients = []
    max_nutrients = []
    wb = openpyxl.load_workbook(std_excel_file, data_only=True)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=3, values_only=True):
        req = float(row[0]) if row[0] is not None else 0.0
        max_val = float(row[1]) if row[1] is not None else float("inf")
        required_nutrients.append(req)
        max_nutrients.append(max_val)
    wb.close()

    # ★ 複数回試行（例：5回）
    trial_results = {}
    num_trials = 5
    seen_signatures = set()  # 既に出現した結果の署名を保持

    for trial in range(1, num_trials+1):
        # modeに応じた制約値の調整
        modified_required_nutrients = required_nutrients[:]  # コピー
        if mode == 'normal':
            modified_required_nutrients = [val * random.uniform(0.75, 1.25) for val in required_nutrients]
        elif mode == 'tanpaku':
            modified_required_nutrients[1] *= random.uniform(1.5, 2.0)
        elif mode == 'karori':
            modified_required_nutrients[0] *= random.uniform(1.5, 2.0)
        else:
            raise ValueError(f"未知のモード: {mode}")

        # 最適化問題の定義
        problem = pulp.LpProblem(f"Menu_Optimization_Trial_{trial}", pulp.LpMinimize)
        # 各料理の選択数（重複選択可、整数変数）
        x = pulp.LpVariable.dicts("x", menus, lowBound=0, cat='Integer')

        # 目的関数：総価格の最小化
        problem += pulp.lpSum([prices[m] * x[m] for m in menus])

        # 各栄養素の制約
        for i, nutrient_name in enumerate(nutrient_names):
            problem += pulp.lpSum([nutrients[m][i] * x[m] for m in menus]) >= modified_required_nutrients[i], f"{nutrient_name}_min_constraint_{trial}"

        problem.solve(pulp.PULP_CBC_CMD(msg=False))

        # 試行結果の集計：数量が1以上の料理のみ
        current_result = {}
        for m in menus:
            quantity = int(pulp.value(x[m]))
            if quantity >= 1:
                current_result[m] = {
                    "quantity": quantity,
                    "image_url": image_urls[m],
                    "price": prices[m]
                }
        
        # 現在の試行結果の署名を作成（料理名と数量の組み合わせで判定）
        signature = tuple(sorted((m, current_result[m]["quantity"]) for m in current_result))
        # 既出の結果と同じ場合は追加しない
        if signature not in seen_signatures:
            trial_results[f"trial_{trial}"] = current_result
            seen_signatures.add(signature)

        #trial_results[f"trial_{trial}"] = current_result

    time.sleep(3)
    return trial_results

# 単体テスト用
if __name__ == '__main__':
    result = get_optimized_menu("カロリー", url="http://example.com")
    print("最適化結果:")
    for trial_key, trial_data in result.items():
        print(trial_key)
        for menu, info in trial_data.items():
            print(f"  {menu}: {info['quantity']}個, {info['price']}円, 画像URL: {info['image_url']}")


# 単体テスト用
if __name__ == '__main__':
    result = get_optimized_menu("カロリー", url="http://example.com")
    print("最適化結果:")
    for trial_key, trial_data in result.items():
        print(trial_key)
        for menu, info in trial_data.items():
            print(f"  {menu}: {info['quantity']}個, {info['price']}円, 画像URL: {info['image_url']}")
